"""Text Collection — save important statements from papers with their source.

When the user selects/highlights text, a snippet is saved together with the
page number, paper title, and source file, so a literature-review note stays
connected to where it came from. Can jump back to the exact spot, export to
Excel, and reopen later. No AI, fully offline.
"""

import re


def _norm_key(text: str) -> str:
    return re.sub(r"[^a-z0-9]", "", (text or "").lower())[:120]


class TextCollection:
    """Holds saved text snippets in memory, with de-duplication."""

    COLUMNS = ["Snippet", "Paper title", "Author", "Page", "Source file",
               "Reference", "Date added"]

    def __init__(self):
        self.entries = []
        self._keys = set()

    def __len__(self):
        return len(self.entries)

    def add(self, snippet: str, paper_title: str = "", author: str = "",
            page: int = 0, source_file: str = "", reference: str = "",
            scroll_pos: int = 0) -> bool:
        """Save one snippet. Returns True if added, False if duplicate.
        `page` is 1-based for display; `scroll_pos` lets us jump back."""
        snippet = " ".join((snippet or "").split()).strip()
        if not snippet:
            return False
        key = _norm_key(snippet)
        if not key or key in self._keys:
            return False
        self._keys.add(key)
        from datetime import datetime
        self.entries.append({
            "Snippet": snippet,
            "Paper title": paper_title,
            "Author": author,
            "Page": page,
            "Source file": source_file,
            "Reference": reference,
            "Date added": datetime.now().strftime("%Y-%m-%d %H:%M"),
            # internal, not exported as a normal column but kept for jump-back
            "_scroll": int(scroll_pos),
        })
        return True

    def remove_at(self, index: int):
        if 0 <= index < len(self.entries):
            e = self.entries.pop(index)
            self._keys.discard(_norm_key(e["Snippet"]))

    def clear(self):
        self.entries.clear()
        self._keys.clear()

    def search(self, term: str) -> list:
        term = (term or "").strip().lower()
        if not term:
            return list(self.entries)
        out = []
        for e in self.entries:
            blob = " ".join(str(e.get(c, "")) for c in self.COLUMNS).lower()
            if term in blob:
                out.append(e)
        return out

    # ---- Excel ----
    def save_xlsx(self, path: str):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        wb = Workbook()
        ws = wb.active
        ws.title = "Notes"
        ws.append(self.COLUMNS)
        head_fill = PatternFill("solid", start_color="2667FF")
        for col in range(1, len(self.COLUMNS) + 1):
            c = ws.cell(row=1, column=col)
            c.font = Font(bold=True, color="FFFFFF", name="Arial")
            c.fill = head_fill
            c.alignment = Alignment(vertical="center")
        for e in self.entries:
            ws.append([e.get(col, "") for col in self.COLUMNS])
        widths = [60, 30, 22, 7, 24, 50, 18]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        # wrap the snippet column
        for row in range(2, len(self.entries) + 2):
            ws.cell(row=row, column=1).alignment = Alignment(
                wrap_text=True, vertical="top")
        ws.freeze_panes = "A2"
        wb.save(path)

    def load_xlsx(self, path: str, merge: bool = True):
        from openpyxl import load_workbook
        if not merge:
            self.clear()
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return 0
        header = [str(h).strip() if h else "" for h in rows[0]]
        idx = {h: i for i, h in enumerate(header)}
        s_i = idx.get("Snippet", 0)
        added = 0
        for r in rows[1:]:
            if not r or s_i >= len(r) or not r[s_i]:
                continue
            def g(name):
                i = idx.get(name)
                return (str(r[i]) if i is not None and i < len(r) and r[i]
                        else "")
            page_val = g("Page")
            try:
                page_num = int(float(page_val)) if page_val else 0
            except Exception:
                page_num = 0
            if self.add(str(r[s_i]), g("Paper title"), g("Author"),
                        page_num, g("Source file"), g("Reference")):
                added += 1
        return added
