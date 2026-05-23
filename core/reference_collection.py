"""Reference collection — build a literature-review database while reading.

As the user clicks citations/references in papers, entries are collected here,
de-duplicated, and can be saved to / loaded from an Excel (.xlsx) file so a
researcher can grow one collection across many papers. No AI, fully offline.
"""

import re
import os


# Try to pull a year and authors out of a raw reference string, so the Excel
# has useful columns. This is best-effort plain pattern matching (no AI).
_YEAR = re.compile(r"\b(19|20)\d{2}[a-z]?\b")


def _guess_year(text: str) -> str:
    years = _YEAR.findall(text)
    # findall returns the group; re-search for the full match
    m = list(re.finditer(r"\b((?:19|20)\d{2})[a-z]?\b", text))
    return m[-1].group(1) if m else ""


def _guess_authors(text: str) -> str:
    """Take the part before the year as a rough 'authors' guess."""
    m = re.search(r"\b(?:19|20)\d{2}\b", text)
    head = text[:m.start()] if m else text[:60]
    # strip a leading number marker like "[12]" or "12."
    head = re.sub(r"^\s*(\[\d+\]|\d+\.)\s*", "", head)
    return head.strip(" ,.;").strip()


def _guess_title(text: str) -> str:
    """Rough title guess: the chunk after the year, first sentence-ish."""
    m = re.search(r"\b(?:19|20)\d{2}[a-z]?\b", text)
    tail = text[m.end():] if m else text
    tail = tail.strip(" ,.;:")
    # take up to the first period that ends a title-like chunk
    parts = re.split(r"\.\s", tail, maxsplit=1)
    return parts[0].strip() if parts else tail[:120]


def _norm_key(text: str) -> str:
    """A normalized key for duplicate detection: lowercase, letters/numbers
    only, collapsed. Two references that are basically the same map to the
    same key even if spacing/punctuation differs."""
    return re.sub(r"[^a-z0-9]", "", text.lower())


# DOI pattern, e.g. 10.1038/s41586-020-1234-5
_DOI = re.compile(r"\b10\.\d{4,9}/[-._;()/:A-Za-z0-9]+\b")


def find_doi(text: str) -> str:
    """Return the first DOI found in the text, or ''."""
    m = _DOI.search(text or "")
    return m.group(0).rstrip(".,;)") if m else ""


def scholar_url(entry: dict) -> str:
    """Build the best Google Scholar search URL for a reference entry.

    Prefers a DOI (most precise), then title, then the full reference text.
    """
    from urllib.parse import quote_plus
    full = entry.get("Full reference", "") if isinstance(entry, dict) else str(entry)
    doi = find_doi(full)
    if doi:
        query = doi
    else:
        title = entry.get("Title", "") if isinstance(entry, dict) else ""
        author = entry.get("Authors", "") if isinstance(entry, dict) else ""
        if title and len(title) > 5:
            query = title
            if author:
                # add first author surname to sharpen the search
                first = author.split(",")[0].split()[-1] if author.split() else ""
                if first:
                    query = f"{title} {first}"
        else:
            query = full[:200]
    return "https://scholar.google.com/scholar?q=" + quote_plus(query)


class ReferenceCollection:
    """Holds collected references in memory, with de-duplication."""

    COLUMNS = ["Number", "Authors", "Year", "Title", "Full reference",
               "Source paper", "Date added"]

    def __init__(self):
        self.entries = []          # list of dicts (the COLUMNS keys)
        self._keys = set()         # normalized keys for dedup

    def __len__(self):
        return len(self.entries)

    def add(self, full_reference: str, source_paper: str = "") -> bool:
        """Add one reference. Returns True if added, False if it was a
        duplicate (and therefore skipped)."""
        full_reference = " ".join(full_reference.split()).strip()
        if not full_reference:
            return False
        key = _norm_key(full_reference)
        if not key or key in self._keys:
            return False
        self._keys.add(key)
        from datetime import datetime
        # pull a leading number like "[12]" if present
        mnum = re.match(r"^\s*\[?(\d{1,3})\]?[.)]?\s", full_reference)
        number = mnum.group(1) if mnum else str(len(self.entries) + 1)
        self.entries.append({
            "Number": number,
            "Authors": _guess_authors(full_reference),
            "Year": _guess_year(full_reference),
            "Title": _guess_title(full_reference),
            "Full reference": full_reference,
            "Source paper": source_paper,
            "Date added": datetime.now().strftime("%Y-%m-%d %H:%M"),
        })
        return True

    def add_many(self, refs: list, source_paper: str = "") -> int:
        """Add a list of reference strings. Returns how many were newly added."""
        added = 0
        for r in refs:
            if self.add(r, source_paper):
                added += 1
        return added

    def remove_at(self, index: int):
        if 0 <= index < len(self.entries):
            e = self.entries.pop(index)
            self._keys.discard(_norm_key(e["Full reference"]))

    def clear(self):
        self.entries.clear()
        self._keys.clear()

    def search(self, term: str) -> list:
        """Return entries whose text contains the term (case-insensitive)."""
        term = (term or "").strip().lower()
        if not term:
            return list(self.entries)
        out = []
        for e in self.entries:
            blob = " ".join(str(v) for v in e.values()).lower()
            if term in blob:
                out.append(e)
        return out

    # ---- Excel save / load ----
    def save_xlsx(self, path: str):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = Workbook()
        ws = wb.active
        ws.title = "References"
        # header row
        ws.append(self.COLUMNS)
        head_fill = PatternFill("solid", start_color="2667FF")
        for col, _ in enumerate(self.COLUMNS, start=1):
            c = ws.cell(row=1, column=col)
            c.font = Font(bold=True, color="FFFFFF", name="Arial")
            c.fill = head_fill
            c.alignment = Alignment(vertical="center")
        # data rows
        for e in self.entries:
            ws.append([e.get(col, "") for col in self.COLUMNS])
        # column widths
        widths = [9, 26, 7, 40, 60, 24, 18]
        from openpyxl.utils import get_column_letter
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        wb.save(path)

    def load_xlsx(self, path: str, merge: bool = True):
        """Load references from an Excel file. If merge=True, add to the
        current collection (skipping duplicates); else replace."""
        from openpyxl import load_workbook
        if not merge:
            self.clear()
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return 0
        header = [str(h).strip() if h else "" for h in rows[0]]
        # map header -> index
        idx = {h: i for i, h in enumerate(header)}
        full_i = idx.get("Full reference",
                         len(header) - 3 if len(header) >= 3 else 0)
        src_i = idx.get("Source paper", None)
        added = 0
        for r in rows[1:]:
            if not r:
                continue
            full = r[full_i] if full_i < len(r) and r[full_i] else ""
            src = (r[src_i] if src_i is not None and src_i < len(r) and r[src_i]
                   else "")
            if full and self.add(str(full), str(src)):
                added += 1
        return added
